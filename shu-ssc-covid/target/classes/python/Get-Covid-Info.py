from inspect import _void
import requests
from lxml import etree
import json
import openpyxl
import time
import pymysql


def GetRowData() -> dict:
    url = "https://voice.baidu.com/act/newpneumonia/newpneumonia#tab"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36"}
    response = requests.get(url, headers=headers)
    html = etree.HTML(response.text)
    result = html.xpath("//*[@id=\"captain-config\"]/text()")
    return json.loads(result[0])["component"][0]


def DealTime(intTime: str) -> str:
    if len(intTime) == 10:
        tupTime = time.localtime(int(intTime))
    elif len(intTime) == 13:
        tupTime = time.localtime(float(int(intTime) / 1000))
    stadardTime = time.strftime("%Y-%m-%d %H:%M:%S", tupTime)
    return stadardTime


def GetSumDomData(result: dict) -> list:
    result = result["summaryDataIn"]
    sumDomData = {}
    sumDomKey = {
        "confirmed": "confirmed",
        "died": "died",
        "cured": "cured",
        "asymptomatic": "asymptomatic",
        "asymptomaticRelative": "asymptomatic_relative",
        "unconfirmed": "unconfirmed",
        "relativeTime": "relative_time",
        "confirmedRelative": "confirmed_relative",
        "unconfirmedRelative": "unconfirmed_relative",
        "curedRelative": "cured_relative",
        "diedRelative": "died_relative",
        "icu": "serious",
        "icuRelative": "serious_relative",
        "overseasInput": "foreign_input",
        "unOverseasInputCumulative": "118604",
        "overseasInputRelative": "foreign_input_relative",
        "unOverseasInputNewAdd": "91",
        "curConfirm": "cur_confirm",
        "curConfirmRelative": "cur_confirm_relative",
        "icuDisable": "1"
    }
    for key in result:
        if key in sumDomKey and not sumDomKey[key].isdigit():
            if key == "relativeTime":
                sumDomData[sumDomKey[key]] = DealTime(result[key])
            else:
                sumDomData[sumDomKey[key]] = result[key]
        elif key in sumDomKey and not sumDomKey[key].isdigit():
            sumDomData[key] = result[key]
    temp = []
    temp.append(sumDomData)
    return temp


def GetProvinceData(result: dict) -> list:
    temp = result["caseList"]
    provinceKey = {
        "confirmed": "confirmed",
        "died": "died",
        "crued": "cured",
        "relativeTime": "relative_time",
        "confirmedRelative": "confirm_relative",
        "diedRelative": "died_relative",
        "curedRelative": "cured_relative",
        "asymptomaticRelative": "asymptomatic_relative",
        "asymptomatic": "asymptomatic",
        "nativeRelative": "0",
        "curConfirm": "cur_confirm",
        "curConfirmRelative": "cur_confirm_relative",
        "overseasInputRelative": "0",
        "icuDisable": "1",
        "area": "province"
    }
    for item in temp:
        item.pop("subList")
    provinceData = []
    for i in range(len(temp)):
        provinceData.append({})
        for key in temp[i]:
            if key in provinceKey and not provinceKey[key].isdigit():
                provinceData[i][provinceKey[key]] = DealTime(temp[i][key]) if key == "relativeTime" else temp[i][key]
            elif key in provinceKey and not provinceKey[key].isdigit():
                provinceData[i][key] = temp[i][key]
    return provinceData


def GetNews() -> list:
    url = "https://sa.sogou.com/new-weball/page/sgs/epidemic?type_page=VR"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36"}
    response = requests.get(url, headers=headers)
    response = response.text.encode("ISO-8859-1").decode(requests.utils.get_encodings_from_content(response.text)[0])
    html = etree.HTML(response)
    target = html.xpath("/html/body/script[1]/text()")
    rawNewsData: str = "}," + target[0].split("timeline")[2].split("columns")[0].split("[")[1].split("]")[0] + ",{"
    rawNewsData: list = rawNewsData.split("},{")
    rawNewsData.pop(0)
    rawNewsData.pop()
    for i in range(len(rawNewsData)):
        rawNewsData[i] = "{" + rawNewsData[i]
        rawNewsData[i] += "}"
        rawNewsData[i] = json.loads(rawNewsData[i])
    newsKey = {"timestamp": "time", "url": "link", "title": "title", "content": "content", "source": "source"}
    newsData = []
    for i in range(len(rawNewsData)):
        newsData.append({})
        for key in rawNewsData[i]:
            if key in newsKey and not newsKey[key].isdigit():
                newsData[i][newsKey[key]] = DealTime(str(rawNewsData[i][key])) if key == "timestamp" else rawNewsData[i][key].strip()
            elif key in newsKey and not newsKey[key].isdigit():
                newsData[i][key] = rawNewsData[i][key].strip()
    return newsData


def SaveResult(data: list, method: str, fileName: str = "", sheetName: str = "") -> _void:
    if method == "Console":
        for item in data:
            for key in item:
                print(key + "：" + item[key])
            print()
        print()
    elif method == "Excel":
        if ".xlsx" not in fileName:
            fileName += ".xlsx"
        try:
            file = openpyxl.load_workbook(fileName)
        except:
            file = openpyxl.Workbook()
            file.remove(file["Sheet"])
        if sheetName in file.sheetnames:
            file.remove(file[sheetName])
        sheet = file.create_sheet(sheetName)
        sheet.append(list(data[0].keys()))
        for item in data:
            sheet.append(list(item.values()))
        file.save(fileName)
    elif method == "Text":
        if ".txt" not in fileName:
            fileName += ".txt"
        with open(fileName, "w+", encoding="utf-8") as file_object:
            for item in data:
                for key in item:
                    file_object.write(key + "：" + item[key] + "\n")
                file_object.write("\n")
            file_object.write("\n")
    elif method == "MySQL":
        connection = pymysql.connect(host="localhost", user="root", password="Vurtne1012", charset="utf8")
        cursor = connection.cursor()
        try:
            sql = "CREATE DATABASE IF NOT EXISTS covid;"
            cursor.execute(sql)
            cursor.close()
            connection.close()
            connection = pymysql.connect(host="localhost", user="root", password="Vurtne1012", database="covid", charset="utf8")
            cursor = connection.cursor()
            sql = """DROP TABLE IF EXISTS """ + fileName + """;"""
            cursor.execute(sql)
            sql = "CREATE TABLE " + fileName + " (id INT PRIMARY KEY,"
            count = 1
            for key in data[0]:
                if key == "relative_time" or key == "province" or key == "time":
                    sql += key + " CHAR(20)"
                elif key == "title":
                    sql += key + " CHAR(100)"
                elif key == "content" or key == "link" or key == "source":
                    sql += key + " TEXT"
                else:
                    sql += key + " INT"
                if count == len(data[0]):
                    sql += ");"
                else:
                    sql += ","
                count += 1
            cursor.execute(sql)
            for i in range(len(data)):
                count = 1
                sql = "INSERT INTO " + fileName + " VALUES (" + str(i + 1) + ","
                for key in data[i]:
                    if key == "relative_time" or key == "province" or key == "time" or key == "title" or key == "content" or key == "link" or key == "source":
                        sql += "\"" + data[i][key] + "\""
                    else:
                        sql += data[i][key]
                    if count == len(data[i]):
                        sql += ");"
                    else:
                        sql += ","
                    count += 1
                cursor.execute(sql)
                connection.commit()
            cursor.close()
            connection.close()
        except Exception:
            print(Exception)


rowResult = GetRowData()
sumDomData = GetSumDomData(rowResult)
SaveResult(sumDomData, "MySQL", "sumdom")
provinceData = GetProvinceData(rowResult)
SaveResult(provinceData, "MySQL", "province")
newsData = GetNews()
SaveResult(newsData, "MySQL", "news")